# -*- coding: utf-8 -*-
"""演示用例：操作类型分类识别 — POM / 传统 / AI 对比

两个用例:
  用例1: POM风格登录(展开为逐个关键字步骤) + 传统断言
  用例2: AI操作 + AI断言 + AI执行 混合

颜色标识 (Excel中直观可见):
  蓝色  — 传统操作 (ACTION)
  橙色  — 断言 (ASSERTION / AI_ASSERTION)
  绿色  — AI操作 (AI_ATOMIC / AI_COMPOSITE)

生成后:
  python main.py --type=excel --cases=./examples/reelmate-cases-excel
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
AI_FILL = PatternFill("solid", fgColor="E2EFDA")
TRAD_FILL = PatternFill("solid", fgColor="DAEEF3")
ASSERT_FILL = PatternFill("solid", fgColor="FDE9D9")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

CASE_HEADERS = ["用例编号", "模块", "功能", "用例标题", "步骤", "测试步骤",
                "操作类型", "数据内容", "用例类型"]
CASE_COL_WIDTHS = [14, 20, 14, 38, 8, 28, 22, 68, 10]

# ── 元素定位器 (放在 context.xlsx _WEB页面元素 sheet 中) ──
# 用户名输入框: role=textbox, name=username
# 密码输入框: role=textbox, name=password
# 登录按钮: role=button, name=登录
# 创作按钮: role=button, name=创作


def category_fill(action_type: str) -> PatternFill:
    if action_type.startswith("AI:操作") or action_type.startswith("AI:执行"):
        return AI_FILL
    if action_type.startswith("AI:断言") or (
        action_type.startswith("断言") and not action_type.startswith("AI:")
    ):
        return ASSERT_FILL
    return TRAD_FILL


def write_header(ws):
    for col_idx, h in enumerate(CASE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN


def write_case_rows(ws, start_row: int, case_id: str, module: str, feature: str,
                    title: str, steps: list[tuple[str, str, str]]) -> int:
    """写入一个用例的多行步骤。返回下一可用行号。"""
    for i, (step_desc, action_type, data_content) in enumerate(steps):
        row = start_row + i
        fill = category_fill(action_type)

        if i == 0:
            ws.cell(row=row, column=1, value=case_id)
            ws.cell(row=row, column=2, value=module)
            ws.cell(row=row, column=3, value=feature)
            ws.cell(row=row, column=4, value=title)
            ws.cell(row=row, column=9, value="WebCase")

        ws.cell(row=row, column=5, value=i + 1)
        ws.cell(row=row, column=6, value=step_desc)
        ws.cell(row=row, column=7, value=action_type)
        ws.cell(row=row, column=8, value=data_content)

        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = fill

    return start_row + len(steps)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "操作类型分类演示"
    write_header(ws)

    row = 2  # 当前数据行

    # ════════════════════════════════════════════════════════════
    # 用例1: POM风格 — 登录流程展开为逐个关键字步骤
    # ════════════════════════════════════════════════════════════
    login_steps = [
        # ── 打开登录页 ──
        ("打开首页", "访问网址",
         "网址={{base_url}}"),
        ("等待页面渲染完成", "强制等待",
         "数据内容=3"),

        # ── 填写表单 ──
        ("点击用户名输入框", "点击元素",
         "定位器类型=role\n角色=textbox\n名称=username"),
        ("输入用户名", "输入内容",
         "定位器类型=role\n角色=textbox\n名称=username\n数据内容={{username}}"),
        ("点击密码输入框", "点击元素",
         "定位器类型=role\n角色=textbox\n名称=password"),
        ("输入密码", "输入内容",
         "定位器类型=role\n角色=textbox\n名称=password\n数据内容={{password}}"),

        # ── 提交 ──
        ("点击登录按钮", "点击元素",
         "定位器类型=role\n角色=button\n名称=登录"),
        ("等待登录结果", "强制等待",
         "数据内容=5"),

        # ── 断言验证 ──
        ("验证跳转到首页", "断言浏览器路径",
         "数据内容=/home"),
        ("验证创作按钮可见", "断言元素存在",
         "定位器类型=role\n角色=button\n名称=创作"),
    ]
    row = write_case_rows(ws, row,
        case_id="DEMO-001",
        module="登录模块",
        feature="POM风格登录",
        title="POM展开-正确账号密码登录成功",
        steps=login_steps)

    # ════════════════════════════════════════════════════════════
    # 用例2: AI混合 — AI原子操作 + AI断言 + AI组合操作
    # ════════════════════════════════════════════════════════════
    ai_steps = [
        # ── 先用传统方式登录 ──
        ("打开首页", "访问网址",
         "网址={{base_url}}"),
        ("等待页面加载", "强制等待",
         "数据内容=3"),
        ("点击登录入口", "点击元素",
         "定位器类型=role\n角色=link\n名称=登录"),
        ("输入账号密码登录", "输入内容",
         "定位器类型=role\n角色=textbox\n名称=username\n数据内容={{username}}"),
        ("等待登录跳转", "强制等待",
         "数据内容=5"),

        # ── 传统断言确认登录成功 ──
        ("验证已登录", "断言元素存在",
         "定位器类型=role\n角色=button\n名称=创作"),

        # ── AI原子操作: 视觉定位点击 ──
        ("AI视觉定位点击参考生视频入口", "AI:操作",
         "操作描述=点击页面中的参考生视频入口按钮"),

        # ── AI断言: 视觉判断 ──
        ("AI验证视频模型下拉框有Seedance选项", "AI:断言",
         "操作描述=视频模型下拉框中存在Seedance 2.0VIP选项"),

        # ── AI组合操作: 多轮Agent ──
        ("AI执行多宫格+TGI2选择流程", "AI:执行",
         "操作描述=选择多宫格模式，然后选择TGI2模型\n最大步数=3"),

        # ── 传统方式收尾 ──
        ("获取最终URL", "获取当前URL",
         "变量名=final_url"),
        ("验证流程完成", "断言浏览器路径",
         "数据内容=video"),
    ]
    row = write_case_rows(ws, row,
        case_id="DEMO-002",
        module="生视频模块",
        feature="AI驱动混合",
        title="AI混合-传统登录+AI操作+AI断言+AI组合执行",
        steps=ai_steps)

    max_row = row - 1

    # ── 列宽 ──
    for i, w in enumerate(CASE_COL_WIDTHS):
        ws.column_dimensions[chr(65 + i)].width = w

    # ── 边框 + 对齐 ──
    for r in ws.iter_rows(min_row=1, max_row=max_row, max_col=9):
        for cell in r:
            cell.border = THIN_BORDER
            if cell.row > 1:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── 图例行 ──
    legend_start = max_row + 2
    legends = [
        (TRAD_FILL, "蓝色 — 传统操作 (ACTION)      例: 访问网址 / 点击元素 / 强制等待 / 输入内容"),
        (ASSERT_FILL, "橙色 — 断言 (ASSERTION)      例: 断言浏览器路径 / 断言元素存在 / AI:断言"),
        (AI_FILL, "绿色 — AI操作 (AI_ATOMIC / AI_COMPOSITE)  例: AI:操作 / AI:执行"),
    ]
    for i, (fill, desc) in enumerate(legends):
        r = legend_start + i
        ws.cell(row=r, column=1, value="■").fill = fill
        ws.cell(row=r, column=2, value=desc)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

    path = os.path.join(OUT_DIR, "0_AIvs传统操作对比演示.xlsx")
    wb.save(path)
    print(f"[OK] Generated: {path}")
    print()
    print("两个用例:")
    print("  DEMO-001: POM展开-登录流程 (关键字逐步骤, 蓝色+橙色)")
    print("  DEMO-002: AI混合 (传统→AI操作→AI断言→AI:执行, 蓝+橙+绿)")
    print()
    print("颜色标注:")
    print("  蓝色 — 传统操作")
    print("  橙色 — 断言")
    print("  绿色 — AI操作")
    print()
    print("验证:")
    print("  python main.py --list-operations")


if __name__ == "__main__":
    main()
