# -*- coding: utf-8 -*-
"""
自动生成 reelmate.cn (万兴剧厂) 登录模块 Excel 测试用例

Excel 用例格式（与参考格式一致）:
  列: 用例编号 | 模块 | 功能 | 用例标题 | 步骤 | 测试步骤 | 操作类型 | 数据内容 | 用例类型
  - 用例编号: 唯一标识（如 DSW-1001），用于 Allure 报告关联
  - 步骤: 步骤序号（1, 2, 3...）
  - 数据内容: key='value' 格式，多个参数用 \\n 分隔
  - 后续行: 用例编号/模块/功能/用例标题/用例类型 留空，只填步骤相关列

数据驱动(DDT):
  Excel 文件可包含第二个 sheet 名为"数据驱动"，格式:
  | 用例标题 | 描述标题 | 参数1 | 参数2 | ... |
  框架会自动将用例中的 {{变量名}} 替换为数据驱动 sheet 中的实际值，
  每行数据展开为一个独立的测试实例。
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

# ---------- 通用样式 ----------
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
DDT_HEADER_FILL = PatternFill("solid", fgColor="7B2D8E")
DDT_DATA_FILL = PatternFill("solid", fgColor="E2EFDA")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 用例表头（与参考格式一致）
CASE_HEADERS = ["用例编号", "模块", "功能", "用例标题", "步骤", "测试步骤", "操作类型", "数据内容", "用例类型"]
CASE_COL_WIDTHS = [14, 20, 14, 28, 8, 22, 22, 60, 10]


def _apply_header_style(ws, headers, fill=HEADER_FILL):
    """写入表头并应用样式。"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = fill
        cell.alignment = HEADER_ALIGN


def _apply_border_and_width(ws, col_widths, max_row, max_col):
    """应用边框和列宽。"""
    for i, w in enumerate(col_widths):
        ws.column_dimensions[chr(65 + i)].width = w
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row > 1:
                cell.alignment = Alignment(vertical="center", wrap_text=True)


def create_context_xlsx():
    """
    创建 context.xlsx 配置文件，包含:
      - 浏览器配置页签
      - WEB页面元素页签（使用 Playwright 原生选择器）
      - 通用配置页签
    """
    wb = Workbook()

    # ===== Sheet 1: 浏览器配置 =====
    ws = wb.active
    ws.title = "浏览器配置"
    ws.append(["浏览器名称", "Grid服务器地址(二选一)", "本地驱动地址(二选一)", "启动参数"])
    ws.append(["chromium", "", "", '{"browserName": "chromium"}'])
    for i, w in enumerate([14, 22, 22, 30], 1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ===== Sheet 2: WEB页面元素 =====
    ws2 = wb.create_sheet("WEB页面元素")
    ws2.append(["元素名称", "定位方式", "目标对象"])
    elements = [
        ["登录入口按钮", "text", "登录"],
        ["用户名输入框", "css",
         'input[type="email"], input[name="email"], input[name="username"], input[type="text"][placeholder*="邮箱"], input[type="text"][placeholder*="账号"]'],
        ["密码输入框", "css", 'input[type="password"]'],
        ["登录提交按钮", "css",
         'button[type="submit"], input[type="submit"], button:has-text("登录"), button:has-text("登錄")'],
        ["登录错误提示", "css", '[class*="error"], [class*="alert"], [class*="message"]'],
        ["用户头像或昵称", "css", '[class*="avatar"], [class*="user"], [class*="nickname"]'],
    ]
    for row in elements:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 16
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 70

    # ===== Sheet 3: 通用配置 =====
    ws3 = wb.create_sheet("通用配置")
    ws3.append(["配置名", "配置值"])
    configs = [
        ["session_reuse", "false"],
        ["username", "18318053665"],
        ["password", "qq111111"],
        ["base_url", "https://www.reelmate.cn"],
        ["login_url", "https://accounts.wondershare.cn/login"],
    ]
    for row in configs:
        ws3.append(row)
    ws3.column_dimensions["A"].width = 16
    ws3.column_dimensions["B"].width = 44

    path = os.path.join(OUT_DIR, "context.xlsx")
    wb.save(path)
    print(f"[OK] 已生成 context.xlsx")


def _write_cases_sheet(ws, case_id_prefix, cases):
    """
    将用例数据写入工作表（参考格式: 用例编号 | 模块 | 功能 | 用例标题 | 步骤 | 测试步骤 | 操作类型 | 数据内容 | 用例类型）。

    首行写完整信息（用例编号、模块、功能、用例标题、用例类型），
    后续行只写步骤相关列（步骤、测试步骤、操作类型、数据内容）。

    :param ws: openpyxl worksheet
    :param case_id_prefix: 用例编号前缀（如 "DSW"）
    :param cases: 用例列表，每项为 (模块, 功能, 用例标题, 用例类型, [(步骤名, 操作类型, 数据内容), ...])
    :return: 写入的最大行号
    """
    row_num = 2
    for case_idx, (module, func, title, case_type, steps) in enumerate(cases):
        case_id = f"{case_id_prefix}-{1001 + case_idx}"
        for step_idx, (step_desc, action_type, data_content) in enumerate(steps):
            # 首行写完整信息，后续行留空
            if step_idx == 0:
                ws.cell(row=row_num, column=1, value=case_id)
                ws.cell(row=row_num, column=2, value=module)
                ws.cell(row=row_num, column=3, value=func)
                ws.cell(row=row_num, column=4, value=title)
                ws.cell(row=row_num, column=9, value=case_type)

            ws.cell(row=row_num, column=5, value=step_idx + 1)
            ws.cell(row=row_num, column=6, value=step_desc)
            ws.cell(row=row_num, column=7, value=action_type)
            ws.cell(row=row_num, column=8, value=data_content)
            row_num += 1

    return row_num - 1


def create_case_xlsx():
    """
    创建登录模块测试用例 Excel (1_登录模块测试.xlsx)

    格式完全参照 1_WEB测试用例.xlsx:
      列: 用例编号 | 模块 | 功能 | 用例标题 | 步骤 | 测试步骤 | 操作类型 | 数据内容 | 用例类型

    数据内容规则:
      - key=value  (值无引号: 数字、URL、元素名等)
      - key='value' (单引号: 含空格/中文的值)
      - 多个参数用 \\n 分隔

    用例列表:
      DSW-1001: 登录页面元素完整性验证
      DSW-1002: 正确账号密码登录成功
      DSW-1003: 错误密码登录失败
      DSW-1004: 空账号校验
      DSW-1005: 空密码校验
      DSW-1006: 格式错误账号校验
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "登录模块测试"
    _apply_header_style(ws, CASE_HEADERS)

    # 数据内容格式与参考文件完全一致:
    #   key=value       — 无引号 (数字、URL、元素名)
    #   key='value'     — 单引号 (中文、含空格的值)
    #   多参数用 \n 分隔
    cases = [
        ("登录模块", "登录页面", "登录页面元素完整性验证", "WebCase", [
            ("打开首页", "访问网址",
             "网址=https://www.reelmate.cn"),
            ("等待页面渲染完成", "强制等待",
             "数据内容=3"),
            ("点击登录入口", "点击元素",
             "_页面元素=登录入口按钮"),
            ("等待登录页加载", "强制等待",
             "数据内容=3"),
            ("验证用户名输入框存在", "断言元素存在",
             "_页面元素=用户名输入框\n超时=10000"),
            ("验证密码输入框存在", "断言元素存在",
             "_页面元素=密码输入框"),
            ("验证登录按钮存在", "断言元素存在",
             "_页面元素=登录提交按钮"),
        ]),
        ("登录模块", "登录功能", "正确账号密码登录成功", "WebCase", [
            ("打开首页", "访问网址",
             "网址=https://www.reelmate.cn"),
            ("等待页面渲染完成", "强制等待",
             "数据内容=3"),
            ("点击登录入口", "点击元素",
             "_页面元素=登录入口按钮"),
            ("等待登录页加载", "强制等待",
             "数据内容=3"),
            ("输入账号", "输入内容",
             "_页面元素=用户名输入框\n数据内容=18318053665"),
            ("输入密码", "输入内容",
             "_页面元素=密码输入框\n数据内容=qq111111"),
            ("点击登录按钮", "点击元素",
             "_页面元素=登录提交按钮"),
            ("等待登录响应与跳转", "强制等待",
             "数据内容=5"),
            ("获取登录后当前URL", "获取当前URL",
             "变量名=after_login_url"),
            ("验证已离开登录页面", "断言文本不相等",
             "预期结果=https://accounts.wondershare.cn/login\n实际结果={{after_login_url}}\n错误信息='登录失败，仍在登录页面'"),
        ]),
        ("登录模块", "登录功能", "错误密码登录失败", "WebCase", [
            ("打开首页", "访问网址",
             "网址=https://www.reelmate.cn"),
            ("等待页面渲染完成", "强制等待",
             "数据内容=3"),
            ("点击登录入口", "点击元素",
             "_页面元素=登录入口按钮"),
            ("等待登录页加载", "强制等待",
             "数据内容=3"),
            ("输入账号", "输入内容",
             "_页面元素=用户名输入框\n数据内容=18318053665"),
            ("输入错误密码", "输入内容",
             "_页面元素=密码输入框\n数据内容=wrongpassword123"),
            ("点击登录按钮", "点击元素",
             "_页面元素=登录提交按钮"),
            ("等待登录响应", "强制等待",
             "数据内容=3"),
            ("获取当前页面URL", "获取当前URL",
             "变量名=current_url"),
            ("验证仍在登录页面", "断言文本包含",
             "预期结果=accounts.wondershare.cn\n实际结果={{current_url}}\n错误信息='错误密码登录不应该成功'"),
        ]),
        ("登录模块", "登录功能", "空账号登录校验", "WebCase", [
            ("打开首页", "访问网址",
             "网址=https://www.reelmate.cn"),
            ("等待页面渲染完成", "强制等待",
             "数据内容=3"),
            ("点击登录入口", "点击元素",
             "_页面元素=登录入口按钮"),
            ("等待登录页加载", "强制等待",
             "数据内容=3"),
            ("清空用户名输入框", "清空输入框",
             "_页面元素=用户名输入框"),
            ("输入密码", "输入内容",
             "_页面元素=密码输入框\n数据内容=qq111111"),
            ("点击登录按钮", "点击元素",
             "_页面元素=登录提交按钮"),
            ("等待响应", "强制等待",
             "数据内容=2"),
            ("获取当前页面URL", "获取当前URL",
             "变量名=current_url"),
            ("验证仍在登录页面", "断言文本包含",
             "预期结果=accounts.wondershare.cn\n实际结果={{current_url}}\n错误信息='空账号登录不应该成功'"),
        ]),
        ("登录模块", "登录功能", "空密码登录校验", "WebCase", [
            ("打开首页", "访问网址",
             "网址=https://www.reelmate.cn"),
            ("等待页面渲染完成", "强制等待",
             "数据内容=3"),
            ("点击登录入口", "点击元素",
             "_页面元素=登录入口按钮"),
            ("等待登录页加载", "强制等待",
             "数据内容=3"),
            ("输入账号", "输入内容",
             "_页面元素=用户名输入框\n数据内容=18318053665"),
            ("清空密码输入框", "清空输入框",
             "_页面元素=密码输入框"),
            ("点击登录按钮", "点击元素",
             "_页面元素=登录提交按钮"),
            ("等待响应", "强制等待",
             "数据内容=2"),
            ("获取当前页面URL", "获取当前URL",
             "变量名=current_url"),
            ("验证仍在登录页面", "断言文本包含",
             "预期结果=accounts.wondershare.cn\n实际结果={{current_url}}\n错误信息='空密码登录不应该成功'"),
        ]),
        ("登录模块", "登录功能", "格式错误账号登录校验", "WebCase", [
            ("打开首页", "访问网址",
             "网址=https://www.reelmate.cn"),
            ("等待页面渲染完成", "强制等待",
             "数据内容=3"),
            ("点击登录入口", "点击元素",
             "_页面元素=登录入口按钮"),
            ("等待登录页加载", "强制等待",
             "数据内容=3"),
            ("输入非法格式账号", "输入内容",
             "_页面元素=用户名输入框\n数据内容=!@#$%^&*()"),
            ("输入密码", "输入内容",
             "_页面元素=密码输入框\n数据内容=qq111111"),
            ("点击登录按钮", "点击元素",
             "_页面元素=登录提交按钮"),
            ("等待响应", "强制等待",
             "数据内容=2"),
            ("获取当前页面URL", "获取当前URL",
             "变量名=current_url"),
            ("验证仍在登录页面", "断言文本包含",
             "预期结果=accounts.wondershare.cn\n实际结果={{current_url}}\n错误信息='非法格式账号登录不应该成功'"),
        ]),
    ]

    max_row = _write_cases_sheet(ws, "DSW", cases)
    _apply_border_and_width(ws, CASE_COL_WIDTHS, max_row, 9)

    path = os.path.join(OUT_DIR, "1_登录模块测试.xlsx")
    wb.save(path)
    print(f"[OK] 已生成 1_登录模块测试.xlsx")


def create_ddt_case_xlsx():
    """
    创建带数据驱动(DDT)的 POM 登录测试用例 Excel (10_POM_DDT_多账号登录测试.xlsx)

    Sheet 1 "测试用例": 用例步骤（POM 模式，使用 {{变量}} 占位符）
    Sheet 2 "数据驱动": 多组测试数据，自动展开为多个测试实例

    数据驱动 sheet 格式:
      | 用例标题 | 描述标题 | username | password | 期望结果 |
      每行数据会将用例步骤中的 {{username}}、{{password}} 等变量替换为实际值，
      生成独立的测试实例（如 DSW-1010-登录测试-正确账号登录成功）。
    """
    wb = Workbook()

    # ===== Sheet 1: 测试用例 =====
    ws = wb.active
    ws.title = "测试用例"
    _apply_header_style(ws, CASE_HEADERS, fill=DDT_HEADER_FILL)

    # 用例步骤中使用 {{变量名}} 占位符，由数据驱动 sheet 提供实际值
    cases = [
        ("登录模块", "POM-DDT登录", "POM多账号数据驱动登录测试", "WebCase", [
            ("导航到登录页面", "LoginPage.navigate_to_login", ""),
            ("输入账号", "LoginPage.enter_username",
             "username={{username}}"),
            ("输入密码", "LoginPage.enter_password",
             "password={{password}}"),
            ("点击登录按钮", "LoginPage.click_login_button", ""),
            ("等待登录响应", "强制等待",
             "数据内容=5"),
            ("获取当前URL", "获取当前URL",
             "变量名=current_url"),
            ("验证登录结果", "断言文本包含",
             "预期结果={{期望结果}}\n实际结果={{current_url}}\n错误信息='登录结果不符合预期'"),
        ]),
    ]

    max_row = _write_cases_sheet(ws, "DSW", cases)
    _apply_border_and_width(ws, CASE_COL_WIDTHS, max_row, 9)

    # ===== Sheet 2: 数据驱动 =====
    ws2 = wb.create_sheet("数据驱动")
    ddt_headers = ["用例标题", "描述标题", "username", "password", "期望结果"]
    _apply_header_style(ws2, ddt_headers, fill=DDT_DATA_FILL)

    # 多组测试数据 —— 每行展开为一个独立测试实例
    ddt_rows = [
        ("POM多账号数据驱动登录测试", "正确账号登录成功", "18318053665", "qq111111", "reelmate.cn"),
        ("POM多账号数据驱动登录测试", "错误密码登录失败", "18318053665", "wrongpassword123", "accounts.wondershare.cn"),
        ("POM多账号数据驱动登录测试", "空账号校验", "", "qq111111", "accounts.wondershare.cn"),
    ]

    for row_idx, (title, desc, user, pwd, expected) in enumerate(ddt_rows, 2):
        ws2.cell(row=row_idx, column=1, value=title)
        ws2.cell(row=row_idx, column=2, value=desc)
        ws2.cell(row=row_idx, column=3, value=user)
        ws2.cell(row=row_idx, column=4, value=pwd)
        ws2.cell(row=row_idx, column=5, value=expected)

    _apply_border_and_width(ws2, [36, 20, 20, 20, 24], len(ddt_rows) + 1, 5)

    path = os.path.join(OUT_DIR, "10_POM_DDT_多账号登录测试.xlsx")
    wb.save(path)
    print(f"[OK] 已生成 10_POM_DDT_多账号登录测试.xlsx (含数据驱动)")


if __name__ == "__main__":
    create_context_xlsx()
    create_case_xlsx()
    create_ddt_case_xlsx()
    print(f"\n运行测试命令:")
    print(f'  python main.py --type=excel --cases="{OUT_DIR}"')
    print(f"\nExcel 用例格式说明:")
    print(f"  列: 用例编号 | 模块 | 功能 | 用例标题 | 步骤 | 测试步骤 | 操作类型 | 数据内容 | 用例类型")
    print(f"  数据内容: key='value' 格式，多个参数用 \\n 分隔")
    print(f"  DDT: Excel 文件中的 '数据驱动' sheet 自动展开为多个测试实例")
    print(f"  用例步骤中使用 {{{{变量名}}}} 占位符，由数据驱动 sheet 提供实际值")
