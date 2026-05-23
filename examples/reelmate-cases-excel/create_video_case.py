# -*- coding: utf-8 -*-
"""Generate the reference video generation test case Excel file.

Hybrid approach:
  - POM (LoginPage) for login — stable, reliable
  - AI:执行 for the video-generation flow — natural language, vision-driven

Test flow:
  Login (POM) → AI executes "选择参考生视频-多宫格模式-TGI2,
  验证视频模型下拉框存在Seedance 2.0VIP"

Usage:  python create_video_case.py
Then:   python main.py --type=excel --cases=.
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

OUT_DIR = os.path.dirname(os.path.abspath(__file__))

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11, name="微软雅黑")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

CASE_HEADERS = ["用例编号", "模块", "功能", "用例标题", "步骤", "测试步骤", "操作类型", "数据内容", "用例类型"]
CASE_COL_WIDTHS = [14, 20, 14, 32, 8, 26, 26, 72, 10]


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "参考生视频测试"

    for col_idx, h in enumerate(CASE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN

    # Hybrid: login via POM (stable), video flow via AI:执行 (natural language)
    steps = [
        ("导航到登录页", "LoginPage.navigate_to_login", ""),
        ("输入账号密码登录", "LoginPage.login",
         "username=18318053665\npassword=qq111111"),
        ("登录后等待跳转", "强制等待", "数据内容=5"),
        ("AI自然语言执行生视频流程", "AI:执行",
         "操作描述=在页面中找到参考生视频入口并点击，选择多宫格模式，选择TGI2模型，验证视频模型下拉框中存在Seedance 2.0VIP选项"),
    ]

    for step_idx, (step_desc, action_type, data_content) in enumerate(steps):
        row = step_idx + 2
        if step_idx == 0:
            ws.cell(row=row, column=1, value="DSW-2001")
            ws.cell(row=row, column=2, value="生视频模块")
            ws.cell(row=row, column=3, value="参考生视频")
            ws.cell(row=row, column=4, value="AI驱动-参考生视频-多宫格-TGI2-Seedance验证")
            ws.cell(row=row, column=9, value="WebCase")

        ws.cell(row=row, column=5, value=step_idx + 1)
        ws.cell(row=row, column=6, value=step_desc)
        ws.cell(row=row, column=7, value=action_type)
        ws.cell(row=row, column=8, value=data_content)

    max_row = len(steps) + 1

    for i, w in enumerate(CASE_COL_WIDTHS):
        ws.column_dimensions[chr(65 + i)].width = w

    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=9):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row > 1:
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    path = os.path.join(OUT_DIR, "11_参考生视频测试.xlsx")
    wb.save(path)
    print(f"[OK] Generated: {path}")
    print()
    print("Run test:")
    print(f'  python main.py --type=excel --cases="{OUT_DIR}"')


if __name__ == "__main__":
    main()
